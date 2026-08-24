from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, DataBarRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter
from pathlib import Path

OUT=Path('WPA_Submission/docs/06_แบบบันทึกคะแนนและสรุปผล_40คน.xlsx')
OUT.parent.mkdir(parents=True,exist_ok=True)
wb=Workbook()
ws=wb.active; ws.title='คะแนนรายบุคคล'
dash=wb.create_sheet('Dashboard')
rub=wb.create_sheet('เกณฑ์คะแนน')
imp=wb.create_sheet('นำเข้า_AR_CSV')

DARK='123B46'; TEAL='0F6B78'; LIGHT='DFF1F4'; WHITE='FFFFFF'; PALE='F8FAFC'
thin=Side(style='thin',color='D1D5DB')

def head(cell,fill=TEAL,size=11):
    cell.fill=PatternFill('solid',fgColor=fill); cell.font=Font(name='Noto Sans Thai',bold=True,color=WHITE,size=size); cell.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True); cell.border=Border(bottom=thin)

def normal(cell,center=False):
    cell.font=Font(name='Noto Sans Thai',size=10); cell.alignment=Alignment(horizontal='center' if center else 'left',vertical='center',wrap_text=True); cell.border=Border(bottom=thin)

# Sheet 1
ws.merge_cells('A1:S1'); ws['A1']='แบบบันทึกคะแนนและสรุปผล AI to ESP32 — ว31101 ม.4/5'; head(ws['A1'],DARK,16)
ws.merge_cells('A2:S2'); ws['A2']='ครูผู้สอน นายศิวัสว์ โตนอก | ภาคเรียนที่ 1 ปีการศึกษา 2569 | นักเรียน 40 คน | 8 กลุ่ม กลุ่มละ 5 คน'; ws['A2'].fill=PatternFill('solid',fgColor=LIGHT); ws['A2'].font=Font(name='Noto Sans Thai',bold=True); ws['A2'].alignment=Alignment(horizontal='center')
headers=['เลขที่','ชื่อ-สกุล','กลุ่ม','Pre /5','Pre %','Post /10','Post %','ผล 70%','พัฒนาการ (จุด%)','AI /15','ESP32 /15','Integration /20','Flow /15','Debug /20','Collaborative /5','ปฏิบัติ /90','รวม /100','Game เสริม /1000','หมายเหตุ']
for c,h in enumerate(headers,1): ws.cell(4,c,h); head(ws.cell(4,c))
for i in range(1,41):
    r=4+i; ws.cell(r,1,i); ws.cell(r,3,(i-1)//5+1)
    ws.cell(r,5,f'=IF(D{r}="","",D{r}/5*100)')
    ws.cell(r,7,f'=IF(F{r}="","",F{r}/10*100)')
    ws.cell(r,8,f'=IF(G{r}="","",IF(G{r}>=70,"ผ่าน","ไม่ผ่าน"))')
    ws.cell(r,9,f'=IF(OR(E{r}="",G{r}=""),"",G{r}-E{r})')
    ws.cell(r,16,f'=IF(COUNTA(J{r}:O{r})=0,"",SUM(J{r}:O{r}))')
    ws.cell(r,17,f'=IF(OR(F{r}="",P{r}=""),"",F{r}+P{r})')
    for c in range(1,20): normal(ws.cell(r,c),center=(c!=2 and c!=19))
for col,w in {'A':7,'B':24,'C':8,'D':9,'E':9,'F':9,'G':9,'H':11,'I':15,'J':9,'K':11,'L':14,'M':9,'N':10,'O':15,'P':12,'Q':12,'R':16,'S':22}.items(): ws.column_dimensions[col].width=w
ws.freeze_panes='A5'; ws.row_dimensions[4].height=38
ws.conditional_formatting.add('H5:H44',CellIsRule(operator='equal',formula=['"ผ่าน"'],fill=PatternFill('solid',fgColor='DCFCE7')))
ws.conditional_formatting.add('H5:H44',CellIsRule(operator='equal',formula=['"ไม่ผ่าน"'],fill=PatternFill('solid',fgColor='FEE2E2')))
ws.conditional_formatting.add('I5:I44',ColorScaleRule(start_type='min',start_color='FEE2E2',mid_type='percentile',mid_value=50,mid_color='FEF3C7',end_type='max',end_color='DCFCE7'))
ws.conditional_formatting.add('Q5:Q44',DataBarRule(start_type='num',start_value=0,end_type='num',end_value=100,color='5B9BD5'))
# validations
for rng,lo,hi in [('D5:D44',0,5),('F5:F44',0,10),('J5:J44',0,15),('K5:K44',0,15),('L5:L44',0,20),('M5:M44',0,15),('N5:N44',0,20),('O5:O44',0,5),('R5:R44',0,1000)]:
    dv=DataValidation(type='decimal',operator='between',formula1=lo,formula2=hi,allow_blank=True); ws.add_data_validation(dv); dv.add(rng)

# Rubric
rub.merge_cells('A1:D1'); rub['A1']='เกณฑ์คะแนนรวม 100 คะแนน'; head(rub['A1'],DARK,15)
for c,h in enumerate(['องค์ประกอบ','คะแนนเต็ม','เกณฑ์ผ่าน/ระดับดี','หลักฐาน'],1): rub.cell(3,c,h); head(rub.cell(3,c))
rows=[('AR Post-test',10,'อย่างน้อย 7/10','ผลเกม AR รายบุคคล'),('AI Station',15,'10/15 ขึ้นไป','การสาธิต Hand/No Hand + คำอธิบาย'),('ESP32 Station',15,'10/15 ขึ้นไป','/on /off + Web Server/Route/GPIO'),('Integration Station',20,'13/20 ขึ้นไป','ระบบ Hand → AI → HTTP → ESP32 → LED'),('System Flow Diagram',15,'11/15 ขึ้นไป','แผนภาพ Data Flow'),('Challenge Debugging',20,'14/20 ขึ้นไป','Diagnosis + Evidence + Test + Solution'),('Collaborative Explanation',5,'3/5 ขึ้นไป','สุ่มถามสมาชิก/Peer Teaching')]
for r,row in enumerate(rows,4):
    for c,v in enumerate(row,1): rub.cell(r,c,v); normal(rub.cell(r,c),center=(c==2))
for col,w in {'A':26,'B':12,'C':24,'D':42}.items(): rub.column_dimensions[col].width=w

# Import CSV
imp.merge_cells('A1:N1'); imp['A1']='พื้นที่นำเข้าผลจากเกม AR'; head(imp['A1'],DARK,14)
imp_headers=['timestamp','name','no','group','pre_score','pre_percent','post_10','post_percent','pass_70','gain_percentage_points','game_1000','first_try_10','attempts','control']
for c,h in enumerate(imp_headers,1): imp.cell(3,c,h); head(imp.cell(3,c))
for c,w in enumerate([22,25,8,8,11,12,11,12,12,20,14,14,12,12],1): imp.column_dimensions[get_column_letter(c)].width=w
imp.freeze_panes='A4'

# Dashboard
dash.merge_cells('A1:J1'); dash['A1']='Dashboard สรุปผลการเรียนรู้ AI to ESP32'; head(dash['A1'],DARK,17)
dash.merge_cells('A2:J2'); dash['A2']='ว31101 ม.4/5 | นายศิวัสว์ โตนอก | เกณฑ์ Post-test ผ่าน 70%'; dash['A2'].fill=PatternFill('solid',fgColor=LIGHT); dash['A2'].font=Font(name='Noto Sans Thai',bold=True); dash['A2'].alignment=Alignment(horizontal='center')
for c,h in enumerate(['นักเรียนที่กรอกชื่อ','ผ่าน Post-test','ไม่ผ่าน','Pre เฉลี่ย %','Post เฉลี่ย %','พัฒนาการเฉลี่ย'],1): dash.cell(4,c,h); head(dash.cell(4,c))
forms=['=COUNTIF(คะแนนรายบุคคล!B5:B44,"<>")','=COUNTIF(คะแนนรายบุคคล!H5:H44,"ผ่าน")','=COUNTIF(คะแนนรายบุคคล!H5:H44,"ไม่ผ่าน")','=IFERROR(AVERAGE(คะแนนรายบุคคล!E5:E44),0','=IFERROR(AVERAGE(คะแนนรายบุคคล!G5:G44),0','=IFERROR(AVERAGE(คะแนนรายบุคคล!I5:I44),0)']
forms[3]='=IFERROR(AVERAGE(คะแนนรายบุคคล!E5:E44),0)'; forms[4]='=IFERROR(AVERAGE(คะแนนรายบุคคล!G5:G44),0)'
for c,f in enumerate(forms,1): dash.cell(5,c,f); normal(dash.cell(5,c),True); dash.cell(5,c).font=Font(name='Noto Sans Thai',bold=True,size=15)
for c,h in enumerate(['กลุ่ม','จำนวนนักเรียน','ผ่าน','Post เฉลี่ย %','คะแนนรวมเฉลี่ย /100'],1): dash.cell(8,c,h); head(dash.cell(8,c))
for g in range(1,9):
    r=8+g; dash.cell(r,1,g); dash.cell(r,2,f'=COUNTIFS(คะแนนรายบุคคล!C5:C44,A{r},คะแนนรายบุคคล!B5:B44,"<>")'); dash.cell(r,3,f'=COUNTIFS(คะแนนรายบุคคล!C5:C44,A{r},คะแนนรายบุคคล!H5:H44,"ผ่าน")'); dash.cell(r,4,f'=IFERROR(AVERAGEIFS(คะแนนรายบุคคล!G5:G44,คะแนนรายบุคคล!C5:C44,A{r}),0)'); dash.cell(r,5,f'=IFERROR(AVERAGEIFS(คะแนนรายบุคคล!Q5:Q44,คะแนนรายบุคคล!C5:C44,A{r}),0)')
    for c in range(1,6): normal(dash.cell(r,c),True)
dash.merge_cells('G4:J4'); dash['G4']='เกณฑ์เป้าหมายสำหรับ วPA'; head(dash['G4'])
dash.merge_cells('G5:J9'); dash['G5']='• นักเรียนอย่างน้อย 32/40 คน ผ่าน Post-test ≥70%\n• นักเรียนอย่างน้อย 30/40 คน มีผลการปฏิบัติระดับดีขึ้นไป\n• ผลงานจริง 90 คะแนน + AR Post-test 10 คะแนน\n• เก็บ Challenge Before/After Feedback เป็นหลักฐานพัฒนาการ'; dash['G5'].fill=PatternFill('solid',fgColor='FFF7D6'); dash['G5'].alignment=Alignment(wrap_text=True,vertical='top'); dash['G5'].font=Font(name='Noto Sans Thai',size=10)
for col,w in {'A':16,'B':17,'C':14,'D':16,'E':20,'F':18,'G':18,'H':18,'I':18,'J':18}.items(): dash.column_dimensions[col].width=w
dash.freeze_panes='A3'
chart=BarChart(); chart.type='col'; chart.style=10; chart.title='Post-test เฉลี่ยรายกลุ่ม'; chart.y_axis.title='เปอร์เซ็นต์'; chart.x_axis.title='กลุ่ม'
data=Reference(dash,min_col=4,min_row=8,max_row=16); cats=Reference(dash,min_col=1,min_row=9,max_row=16); chart.add_data(data,titles_from_data=True); chart.set_categories(cats); chart.height=8; chart.width=13; dash.add_chart(chart,'G11')

wb.save(OUT)
print(f'Saved {OUT}')
